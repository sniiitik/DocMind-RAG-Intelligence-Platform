import json
import requests
from openpyxl import load_workbook

INPUT_XLSX = "RAG_Eval_Experiment_Sheet_completed.xlsx"
OUTPUT_XLSX = "RAG_Eval_Experiment_Sheet_filled_docmind.xlsx"

API_URL = "http://localhost:8000/api/query"

wb = load_workbook(INPUT_XLSX)
ws = wb["Query Log"]

for row in range(4, 54):  # 50 queries: rows 4 to 53
    qnum = ws.cell(row=row, column=1).value
    question = ws.cell(row=row, column=4).value

    if not question:
        continue

    print(f"Running Q{qnum}: {question}")

    try:
        response = requests.post(
            API_URL,
            json={"question": question},
            timeout=120,
        )
        response.raise_for_status()
        data = response.json()

        answer = data.get("answer", "")
        sources = data.get("sources", [])
        context = data.get("context", [])

        context_text = ""

        if isinstance(context, list):
            for i, chunk in enumerate(context, start=1):
                if isinstance(chunk, dict):
                    label = chunk.get("source_label", "")
                    text = chunk.get("text", "")
                    context_text += f"[Chunk {i}] {label}\n{text}\n\n"
                else:
                    context_text += f"[Chunk {i}]\n{str(chunk)}\n\n"
        else:
            context_text = str(context)

        ws.cell(row=row, column=5).value = answer
        ws.cell(row=row, column=6).value = context_text

        # Try to extract percentages from answer object if your backend includes them later.
        # If not, leave G/H blank and fill from Evals manually.
        ws.cell(row=row, column=13).value = "Pass" if "don't have enough information" not in answer.lower() else "Fail"

        if ws.cell(row=row, column=13).value == "Fail":
            ws.cell(row=row, column=14).value = "Automated run: answer said insufficient information."
        else:
            ws.cell(row=row, column=14).value = "Automated run."

    except Exception as e:
        ws.cell(row=row, column=5).value = f"ERROR: {e}"
        ws.cell(row=row, column=13).value = "Fail"
        ws.cell(row=row, column=14).value = "API/script error."

    wb.save(OUTPUT_XLSX)

print(f"Done. Saved to {OUTPUT_XLSX}")


