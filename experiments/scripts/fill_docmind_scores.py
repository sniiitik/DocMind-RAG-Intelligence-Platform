import requests
from openpyxl import load_workbook

INPUT_XLSX = "RAG_Eval_Experiment_Sheet_filled_docmind.xlsx"
OUTPUT_XLSX = "RAG_Eval_Experiment_Sheet_filled_docmind_scores.xlsx"

API_URL = "http://localhost:8000/api/evaluate"

wb = load_workbook(INPUT_XLSX)
ws = wb["Query Log"]

for row in range(4, 54):
    qnum = ws.cell(row=row, column=1).value
    question = ws.cell(row=row, column=4).value
    answer = ws.cell(row=row, column=5).value
    context_text = ws.cell(row=row, column=6).value

    if not question or not answer or not context_text:
        continue

    print(f"Evaluating Q{qnum}")

    payload = {
        "question": question,
        "answer": answer,
        "contexts": [context_text],
    }

    try:
        r = requests.post(API_URL, json=payload, timeout=120)
        r.raise_for_status()
        data = r.json()

        scores = data.get("scores", {})
        ws.cell(row=row, column=7).value = scores.get("faithfulness", "")
        ws.cell(row=row, column=8).value = scores.get("answer_relevancy", "")

    except Exception as e:
        print(f"Error Q{qnum}: {e}")
        ws.cell(row=row, column=7).value = ""
        ws.cell(row=row, column=8).value = ""

    wb.save(OUTPUT_XLSX)

print(f"Done. Saved to {OUTPUT_XLSX}")