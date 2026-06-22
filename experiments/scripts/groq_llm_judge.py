import os, json, re, requests
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv("backend/.env")

GROQ_API_KEY = os.getenv("GROQ_API_KEY")
MODEL = "llama-3.3-70b-versatile"

INPUT_XLSX = "RAG_Eval_Experiment_Sheet_filled_docmind_scores.xlsx"
OUTPUT_XLSX = "RAG_Eval_Experiment_Sheet_with_llm_judge.xlsx"

wb = load_workbook(INPUT_XLSX)
ws = wb["Query Log"]

url = "https://api.groq.com/openai/v1/chat/completions"
headers = {
    "Authorization": f"Bearer {GROQ_API_KEY}",
    "Content-Type": "application/json",
}

def judge(query, context, answer):
    prompt = f"""
You are an evaluator for a RAG system.

QUERY: {query}

RETRIEVED CONTEXT:
{context}

GENERATED ANSWER:
{answer}

Rate the answer on two dimensions:
1. Faithfulness (0-10): Is the answer fully supported by the retrieved context?
2. Relevance (0-10): Does the answer directly address the query?

Return ONLY valid JSON:
{{"faithfulness": X, "relevance": Y}}
"""

    payload = {
        "model": MODEL,
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0,
    }

    r = requests.post(url, headers=headers, json=payload, timeout=120)
    r.raise_for_status()
    text = r.json()["choices"][0]["message"]["content"]

    match = re.search(r"\{.*\}", text, re.DOTALL)
    if not match:
        raise ValueError(f"No JSON found: {text}")

    data = json.loads(match.group(0))
    return data["faithfulness"] / 10, data["relevance"] / 10

for row in range(4, 54):
    qnum = ws.cell(row=row, column=1).value
    query = ws.cell(row=row, column=4).value
    answer = ws.cell(row=row, column=5).value
    context = ws.cell(row=row, column=6).value

    print(f"Judging Q{qnum}")

    try:
        faith, rel = judge(query, context, answer)
        ws.cell(row=row, column=11).value = faith
        ws.cell(row=row, column=12).value = rel
    except Exception as e:
        print(f"Error Q{qnum}: {e}")
        ws.cell(row=row, column=11).value = ""
        ws.cell(row=row, column=12).value = ""

    wb.save(OUTPUT_XLSX)

print(f"Done. Saved to {OUTPUT_XLSX}")