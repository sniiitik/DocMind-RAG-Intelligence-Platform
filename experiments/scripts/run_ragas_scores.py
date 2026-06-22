import os
import pandas as pd
from openpyxl import load_workbook
from datasets import Dataset
from dotenv import load_dotenv

from ragas import evaluate
from ragas.metrics import faithfulness, answer_relevancy
from langchain_openai import ChatOpenAI, OpenAIEmbeddings

load_dotenv("backend/.env")

INPUT_XLSX = "RAG_Eval_Experiment_Sheet_with_llm_judge.xlsx"
OUTPUT_XLSX = "RAG_Eval_Experiment_Sheet_FINAL.xlsx"

# RAGAS expects OpenAI-compatible env vars.
# Groq will be used for the judge LLM.
os.environ["OPENAI_API_KEY"] = os.getenv("GROQ_API_KEY")
os.environ["OPENAI_API_BASE"] = "https://api.groq.com/openai/v1"

llm = ChatOpenAI(
    model="llama-3.3-70b-versatile",
    temperature=0,
    openai_api_key=os.getenv("GROQ_API_KEY"),
    openai_api_base="https://api.groq.com/openai/v1",
)

# Important:
# Groq does NOT provide embeddings, so RAGAS answer_relevancy may fail
# if it needs OpenAI embeddings. If it fails, we will run faithfulness only.
embeddings = None

wb = load_workbook(INPUT_XLSX)
ws = wb["Query Log"]

questions = []
answers = []
contexts = []
rows = []

for row in range(4, 54):
    q = ws.cell(row=row, column=4).value
    a = ws.cell(row=row, column=5).value
    c = ws.cell(row=row, column=6).value

    if not q or not a or not c:
        continue

    rows.append(row)
    questions.append(q)
    answers.append(a)
    contexts.append([c])

dataset = Dataset.from_dict({
    "question": questions,
    "answer": answers,
    "contexts": contexts,
})

print("Running RAGAS...")
print("If answer_relevancy fails because Groq has no embeddings, tell me and we will use a local embedding workaround.")

try:
    results = evaluate(
        dataset,
        metrics=[faithfulness, answer_relevancy],
        llm=llm,
    )

    df = results.to_pandas()

    for idx, row_num in enumerate(rows):
        ws.cell(row=row_num, column=9).value = float(df.loc[idx, "faithfulness"])
        ws.cell(row=row_num, column=10).value = float(df.loc[idx, "answer_relevancy"])

except Exception as e:
    print("RAGAS failed:")
    print(e)
    print("Trying faithfulness only...")

    results = evaluate(
        dataset,
        metrics=[faithfulness],
        llm=llm,
    )

    df = results.to_pandas()

    for idx, row_num in enumerate(rows):
        ws.cell(row=row_num, column=9).value = float(df.loc[idx, "faithfulness"])
        ws.cell(row=row_num, column=10).value = ""

wb.save(OUTPUT_XLSX)
print(f"Done. Saved to {OUTPUT_XLSX}")